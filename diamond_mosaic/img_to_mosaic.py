import svgwrite
from PIL import Image
import numpy as np
from scipy.spatial import cKDTree
import multiprocessing
import concurrent.futures
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.fills import PatternFill

import diamond_mosaic.utils as utils
import diamond_mosaic.config as config


def mosaic_to_pixel(w, h, m_s_cm, pixel_per_centimeter=120):
    """Convert real mosaic size to pixels
    :w: mosaic amount in width
    :h: mosaic amount in height
    :m_s_cm: mosaic diameter in cm.
    :pixel_per_centimeter: default density of image = 120 ppc (304.8 ppi)
    """
    single_mosaic_pixel = pixel_per_centimeter * m_s_cm
    return (round(w * single_mosaic_pixel), round(h * single_mosaic_pixel))


def add_svg_text(svg_img, text_list, mosaic_size):
    coords, mosaic_r_mm = get_ellipse_params(
        len(text_list[0]), len(text_list), mosaic_size
    )
    style = f"text-anchor:middle;stroke-width:{mosaic_r_mm*0.01}mm;stroke:#000000;"
    for i, row in enumerate(coords):
        for j, circle_middle in enumerate(row):
            text_middle_coord = (circle_middle[0], circle_middle[1] + mosaic_r_mm / 3)
            svg_img.add(
                svg_img.text(
                    text_list[i][j],
                    insert=text_middle_coord,
                    fill="white",
                    style=style,
                    font_size=f"{(mosaic_r_mm/3.5)}mm",
                )
            )
    return svg_img


def get_ellipse_coord(pic_size, mosaic_size):
    m_area = list(mosaic_to_pixel(1, 1, mosaic_size))[0]
    num_x = int(pic_size[0] / m_area)
    num_y = int(pic_size[1] / m_area)
    x_coord = [m_area * num for num in range(0, num_x + 1)]
    y_coord = [m_area * num for num in range(0, num_y + 1)]

    coord_up = []
    for y in range(0, num_y):
        coord = [(x_coord[x], y_coord[y]) for x in range(0, num_x)]
        coord_up.append(coord)

    coord_down = []
    for y in range(1, num_y + 1):
        coord = [(x_coord[x], y_coord[y]) for x in range(1, num_x + 1)]
        coord_down.append(coord)
    return coord_up, coord_down


def get_ellipse_params(mosaic_w_amount, mosaic_h_amount, mosaic_size):
    mosaic_d_mm = mosaic_size * 10
    coords = []
    for y in range(1, mosaic_h_amount + 1):
        coords.append(
            [
                (
                    (mosaic_d_mm) * x + mosaic_d_mm / 2,
                    (mosaic_d_mm) * y - mosaic_d_mm / 2,
                )
                for x in range(0, mosaic_w_amount)
            ]
        )
    return coords, mosaic_d_mm / 2


def add_svg_circle(svg_img, color_list, mosaic_size):
    coords, mosaic_r = get_ellipse_params(
        len(color_list[0]), len(color_list), mosaic_size
    )
    for i, row in enumerate(coords):
        for j, circle_middle in enumerate(row):
            color = f"rgb({int(color_list[i][j][0])},{int(color_list[i][j][1])},{int(color_list[i][j][2])})"
            svg_img.add(svg_img.circle(circle_middle, r=mosaic_r, fill=color))
    return svg_img


def close_color(rgb_color, color_data):
    rgb_colors = np.array(list(color_data.values()))
    dmc_code = np.array(list(color_data.keys()))
    similar_color_idx = cKDTree(rgb_colors).query(rgb_color, k=1)[1]
    return rgb_colors[similar_color_idx], dmc_code[similar_color_idx]


def convert_color(input_colors):
    color_palette = utils.read_json(config.DATA_PATH + config.RGB_FILE)
    color_palette = transform_array_in_dict(color_palette)
    converted_color = []
    colors_name = []
    for i in input_colors:
        row_color_rgb = []
        row_color_dmc_code = []

        row_color_rgb, row_color_dmc_code = close_color(i, color_palette)
        converted_color.append(row_color_rgb)
        colors_name.append(row_color_dmc_code)
    return converted_color, colors_name


def transform_array_in_dict(dict_colors):
    for key, value in dict_colors.items():
        dict_colors[key] = np.array(value)
    return dict_colors


def split_weird_array(array):
    for i, value in enumerate(array):
        if i == 0:
            v = value
            if len(array) == 1:
                return np.array(v)
            continue
        ready_array = np.vstack((v, value))
        v = ready_array
    return ready_array


def hold_aspect_ratio(img_size, m_w, m_h):
    if not m_h:
        ratio = img_size[0] / img_size[1]
        if m_w > 2:
            dot_h = int(m_w / ratio)
            dot_w = int(m_w)
        else:
            raise ValueError("Diamonds along width cant be less then <2")
    else:
        dot_h = int(m_h)
        dot_w = int(m_w)
    return dot_w, dot_h


def divide_into_chunks(image):
    cpu_number = multiprocessing.cpu_count()
    img_size = list(image.shape)
    divide = int(img_size[0] / cpu_number)
    chunks = []
    if img_size[0] > cpu_number:
        for val in range(0, cpu_number):
            right_limit = divide * (val + 1)
            left_limit = val * divide
            if val == cpu_number - 1:
                right_limit = img_size[0]
            chunks.append(image[left_limit:right_limit])
    else:
        chunks = [image]
    return chunks


def unique_count(a):
    unique_value, counts = np.unique(a, return_counts=True)
    a_dict = dict(zip(unique_value, counts))
    sort_a = sorted(a_dict.items(), key=lambda item: item[1])[::-1]
    return dict(sort_a)


def save_color_table(file_name, encode_list):
    encode_result = unique_count(encode_list)

    color_file = utils.read_json(config.DATA_PATH + config.HEX_FILE)
    labels_file = utils.read_json(config.DATA_PATH + config.LABELS_FILE)

    encode_colors_list = list(labels_file.values())
    hex_colors_list = list(color_file.values())
    dmc_colors_list = list(color_file.keys())

    sorted_encode = sorted(list(encode_result.keys()))
    sorted_dmc_colors = []
    sorted_hex_colors = []

    for i in sorted_encode:
        for j, val in enumerate(encode_colors_list):
            if i == val:
                sorted_dmc_colors.append(dmc_colors_list[j])
                sorted_hex_colors.append(hex_colors_list[j])

    hex_in_img = []
    dmc_in_img = []
    for i in list(encode_result.keys()):
        for j, val in enumerate(encode_colors_list):
            if i == val:
                hex_in_img.append(hex_colors_list[j])
                dmc_in_img.append(dmc_colors_list[j])

    pd_amount = pd.DataFrame(list(encode_result.values()))
    pd_code = pd.DataFrame(dmc_in_img)
    pd_encode = pd.DataFrame(list(encode_result.keys()))

    df = pd.concat([pd_amount, pd_encode, pd_code], axis=1)
    df.reset_index(drop=True, inplace=True)
    df.columns = ["Amount in pieces", "Color label", "DMC color"]

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for i in range(0, len(hex_in_img)):
        ws["D" + str(i + 2)].fill = PatternFill(
            "solid", start_color=str(hex_in_img[i][1::])
        )

    ws["F1"].value = "In alphabetical order"
    ws["G1"].value = "Color label"
    ws["H1"].value = "DMC color"
    ws["I1"].value = "Color"
    for i in range(0, len(sorted_dmc_colors)):
        ws["G" + str(i + 2)].value = sorted_encode[i]
        ws["H" + str(i + 2)].value = sorted_dmc_colors[i]
        ws["I" + str(i + 2)].fill = PatternFill(
            "solid", start_color=str(sorted_hex_colors[i])[1::]
        )
    wb.save(file_name + ".xlsx")


def get_text_labels(color_in_img):
    labels_data = utils.read_json(config.DATA_PATH + config.LABELS_FILE)
    color_names = list(labels_data.keys())
    labels_values = list(labels_data.values())

    encode_text = []
    for i in color_in_img:
        a_row = [labels_values[color_names.index(pixel)] for pixel in i]
        encode_text.append(a_row)
    return encode_text


def paralell_color_convertion(chunks):
    color_list = []
    color_name = []
    with concurrent.futures.ProcessPoolExecutor() as exe:
        result = exe.map(convert_color, chunks)
        for row_color, row_color_name in result:
            color_list.append(row_color)
            color_name.append(row_color_name)
    color_list = np.array(split_weird_array(color_list))
    color_name = list(split_weird_array(color_name))
    return color_list, color_name


def img_to_mosaic(img_name, mosaic_number_w, mosaic_number_h, mosaic_size, save_name):

    image = Image.open(img_name)
    image = image.convert("RGB")

    mosaic_number_w, mosaic_number_h = hold_aspect_ratio(
        image.size, mosaic_number_w, mosaic_number_h
    )
    resize_img = np.array(image.resize((mosaic_number_w, mosaic_number_h)))

    print(f"Original image size:   {image.size[0]}x{image.size[1]} px")
    mosaic_w = mosaic_size * mosaic_number_w * 10  # mm
    mosaic_h = mosaic_size * mosaic_number_h * 10  # mm

    chunks = divide_into_chunks(resize_img)
    color_list, color_name = paralell_color_convertion(chunks)
    print(
        f"Diamonds picture size: {mosaic_w}x{mosaic_h} mm ({len(color_name[0])}x{len(color_name)} pc)"
    )
    svg_img = svgwrite.Drawing(
        f"{save_name}.svg",
        profile="full",
        size=(
            f"{mosaic_w}mm",
            f"{mosaic_h}mm",
        ),
        viewBox=(f"0 0 {mosaic_w} {mosaic_h}"),
    )
    svg_img = add_svg_circle(svg_img, color_list, mosaic_size)
    encode_text = get_text_labels(color_name)
    svg_img = add_svg_text(svg_img, encode_text, mosaic_size)
    svg_img.save()
    save_color_table(f"{save_name}", encode_text)
