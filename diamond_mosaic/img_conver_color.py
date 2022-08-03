from PIL import Image, ImageDraw, ImageFont
import numpy as np
import time
import json
import multiprocessing
import concurrent.futures
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.fills import PatternFill

from diamond_mosaic import color_finder
from diamond_mosaic import color
from diamond_mosaic import settings


def mosaic_to_pixel(w, h, m_s_cm):
    # 120 = 304.8 dpi
    single_mosaic_pixel = 120 * m_s_cm
    return (round(w * single_mosaic_pixel), round(h * single_mosaic_pixel))


def add_grid(image, m_size):
    black = [0, 0, 0]
    m_area = list(mosaic_to_pixel(1, 1, m_size))[0]
    image_shape = list(image.size)
    num_lines_x = int((image_shape[0] + 1) / m_area)
    num_lines_y = int((image_shape[1] + 1) / m_area)
    lines_x_coord = [num * m_area for num in range(0, num_lines_x + 1)]
    lines_y_coord = [num * m_area for num in range(0, num_lines_y + 1)]

    img = []
    for i, val in enumerate(np.array(image)):
        if i in lines_y_coord:
            a_row = [black for j in val]
        else:
            a_row = [black if j in lines_x_coord else val for j, val in enumerate(val)]
        img.append(a_row)
    return Image.fromarray(np.array(img, dtype=np.uint8))


def get_text_coord(img, m_size):
    m_area = list(mosaic_to_pixel(1, 1, m_size))[0]
    image_shape = list(img.size)

    num_dot_x = int((image_shape[0] + 1) / m_area)
    num_dot_y = int((image_shape[1] + 1) / m_area)

    x_coord = [num * m_area + round(m_area / 2) for num in range(0, num_dot_x)]
    y_coord = [num * m_area + round(m_area / 2) for num in range(0, num_dot_y)]

    coord = []
    for y in y_coord:
        for x in x_coord:
            coord.append((x, y))
    return coord


def add_text(image, text_list, coord_list, m_size):
    draw = ImageDraw.Draw(image)
    m_area = list(mosaic_to_pixel(1, 1, m_size))[0]
    font = ImageFont.truetype(
        "/usr/share/fonts/truetype/quicksand/Quicksand-Regular.ttf",
        int((m_area / 2) + 1),
    )

    unfolded_text = [j for i in text_list for j in i]

    for coord, text in zip(coord_list, unfolded_text):
        draw.text(
            coord,
            text,
            fill="white",
            font=font,
            stroke_width=2,
            stroke_fill="black",
            anchor="mm",
        )
    return image


def get_color_name_in_img(img, color_palate):
    return [[get_color_name(pixel, color_palate) for pixel in i] for i in img]


def get_ellipse_coord(img, mosaic_size):
    m_area = list(mosaic_to_pixel(1, 1, mosaic_size))[0]
    img_size = list(img.size)
    num_x = int(img_size[0] / m_area)
    num_y = int(img_size[1] / m_area)
    x_coord = [m_area * num for num in range(0, num_x + 1)]
    y_coord = [m_area * num for num in range(0, num_y + 1)]

    coord_up = []
    for x in range(0, num_x):
        coord = [(x_coord[x], y_coord[y]) for y in range(0, num_y)]
        coord_up.append(coord)

    coord_down = []
    for y in range(1, num_y + 1):
        coord = [(x_coord[x], y_coord[y]) for x in range(1, num_x + 1)]
        coord_down.append(coord)

    ready_coord_up = [list(i) for i in transpose(coord_up)]
    return ready_coord_up, coord_down


def transpose(matrix):
    return [*zip(*matrix)]


def add_circle(img_plane, up_coord, down_coord, color_list):
    draw = ImageDraw.Draw(img_plane)
    for i, val in enumerate(up_coord):
        for j, x_coord in enumerate(val):
            color = tuple(color_list[i][j])
            draw.ellipse([x_coord, down_coord[i][j]], fill=color)
    return img_plane


def convert_color(input_colors):
    color_palate = color.get_color(settings.PATH + settings.RGB_FILE)
    converted_color = []
    colors_name = []
    finded_unique_colors = {}
    for i in input_colors:
        row = []
        for pixel in i:
            if str(pixel) in finded_unique_colors:
                row.append(finded_unique_colors[str(pixel)])
                continue
            close_color = color_finder.close_color(pixel, color_palate)
            finded_unique_colors[str(pixel)] = close_color
            row.append(close_color)
        row = [i for i in zip(*row)]
        converted_color.append(row[0])
        colors_name.append(row[1])
    return converted_color, colors_name


def split_weird_array(array):
    for i, value in enumerate(array):
        if i == 0:
            v = value
            if len(array) == 1:
                return np.array(v)
            continue
        ready_array = np.vstack((v, value))
        v = ready_array
    return ready_array


def hold_aspect_ratio(img_size, m_w, m_h):
    ratio = img_size[0] / img_size[1]
    max_mosaic = max(m_w, m_h)
    return int(ratio * max_mosaic), int(max_mosaic)


def divide_into_chunks(image):
    cpu_number = multiprocessing.cpu_count()
    img_size = list(image.shape)
    divide = int(img_size[0] / cpu_number)
    chunks = []
    if img_size[0] > cpu_number:
        for val in range(0, cpu_number):
            right_limit = divide * (val + 1)
            left_limit = val * divide
            if val == cpu_number - 1:
                right_limit = img_size[0]
            chunks.append(image[left_limit:right_limit])
    else:
        chunks = [image]
    return chunks


def unique_count(a):
    unique_value, counts = np.unique(a, return_counts=True)
    a_dict = dict(zip(unique_value, counts))
    sort_a = sorted(a_dict.items(), key=lambda item: item[1])[::-1]
    return dict(sort_a)


def save_color_table(file_name, color_list, encode_list):
    color_result = unique_count(color_list)
    encode_result = unique_count(encode_list)

    color_file = Color.get_color(Color.HEX_FILE)
    encode_file = Color.get_color(Color.ENCODE_FILE)

    encode_colors_list = list(encode_file.values())
    hex_colors_list = list(color_file.values())
    dmc_colors_list = list(color_file.keys())

    sorted_encode = sorted(list(encode_result.keys()))
    sorted_dmc_colors = []
    sorted_hex_colors = []

    for i in sorted_encode:
        for j, val in enumerate(encode_colors_list):
            if i == val:
                sorted_dmc_colors.append(dmc_colors_list[j])
                sorted_hex_colors.append(hex_colors_list[j])

    hex_in_img = []
    dmc_in_img = []
    for i in list(encode_result.keys()):
        for j, val in enumerate(encode_colors_list):
            if i == val:
                hex_in_img.append(hex_colors_list[j])
                dmc_in_img.append(dmc_colors_list[j])

    print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

    pd_amount = pd.DataFrame(list(encode_result.values()))
    pd_code = pd.DataFrame(dmc_in_img)
    pd_encode = pd.DataFrame(list(encode_result.keys()))

    df = pd.concat([pd_amount, pd_encode, pd_code], axis=1)
    df.reset_index(drop=True, inplace=True)
    df.columns = ["Кол-во", "Символ", "DMC цвет"]

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    for i in range(0, len(hex_in_img)):
        ws["D" + str(i + 2)].fill = PatternFill(
            "solid", start_color=str(hex_in_img[i][1::])
        )

    ws["F1"].value = "В алфавитном порядке"
    ws["G1"].value = "Символ"
    ws["H1"].value = "DMC цвет"
    ws["I1"].value = "Цвет"
    for i in range(0, len(sorted_dmc_colors)):
        ws["G" + str(i + 2)].value = sorted_encode[i]
        ws["H" + str(i + 2)].value = sorted_dmc_colors[i]
        ws["I" + str(i + 2)].fill = PatternFill(
            "solid", start_color=str(sorted_hex_colors[i])[1::]
        )

    wb.save(file_name + ".xlsx")


def get_encoding_text(color_in_img):
    encoding_data = Color.get_color(Color.ENCODE_FILE)
    color_names = list(encoding_data.keys())
    all_encoding_values = list(encoding_data.values())

    encode_text = []
    for i in color_in_img:
        a_row = [all_encoding_values[color_names.index(pixel)] for pixel in i]
        encode_text.append(a_row)

    return encode_text


def paralell_color_convertion(chunks):
    color_list = []
    color_name = []
    with concurrent.futures.ProcessPoolExecutor() as exe:
        result = exe.map(convert_color, chunks)
        for row_color, row_color_name in result:
            color_list.append(row_color)
            color_name.append(row_color_name)
    color_list = split_weird_array(color_list)
    color_name = list(split_weird_array(color_name))
    return color_list, color_name


def save_img(file_name, img):
    margin = 3  # cm
    margin_top_pix = list(mosaic_to_pixel(1, 1, margin))[0]
    margin_bottom_pix = margin_top_pix
    margin_left_pix = list(mosaic_to_pixel(1, 1, margin))[0]
    margin_right_pix = margin_left_pix

    res = list(mosaic_to_pixel(1, 1, 2.54))[0]
    w, h = img.size
    new_w = w + margin_right_pix + margin_left_pix
    new_h = h + margin_top_pix + margin_bottom_pix

    img_with_pad = Image.new(img.mode, (new_w, new_h), (255, 255, 255))
    img_with_pad.paste(img, (margin_left_pix, margin_top_pix))

    print(f"resolution in PDF = {res} ppi")
    paper_w = (new_w * margin) / margin_top_pix
    paper_h = (new_h * margin) / margin_top_pix

    print(f"paper size width = {paper_w} cm, height = {paper_h} cm.")

    img_with_pad.save(
        file_name + ".pdf",
        "PDF",
        resolution=res,
        author="LALAPOPA",
    )


def img_to_mosaic(img_name, mosaic_number_w, mosaic_number_h):
    mosaic_size = 0.25  # cm

    image = Image.open(img_name)
    color_palate = color.get_color(PATH + RGB_FILE)

    mosaic_number_w, mosaic_number_h = hold_aspect_ratio(
        image.size, mosaic_number_w, mosaic_number_h
    )

    resize_img = image.resize((mosaic_number_w, mosaic_number_h))

    print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    print(f"original size: {image.size} px width x height")
    print(
        f"mosaic picture size width: {mosaic_size*mosaic_number_w} cm, height: {mosaic_size*mosaic_number_h} cm"
    )

    a = np.array(resize_img)
    chunks = divide_into_chunks(a)
    color_list, color_name = paralell_color_convertion(chunks)

    print(f"number of mosaic: {len(color_name[0])}x{len(color_name)}")
    # print(f"Color_list = shape is {color_list.shape}")
    img_color = Image.fromarray(np.array(color_list, dtype=np.uint8))
    img_bg = Image.new("RGB", img_color.size, (255, 255, 255))
    ready_size_img = img_bg.resize(
        mosaic_to_pixel(
            mosaic_number_w,
            mosaic_number_h,
            mosaic_size,
        ),
        resample=Image.BOX,
    )
    up_coord, down_coord = get_ellipse_coord(ready_size_img, mosaic_size)

    ready_img = add_circle(ready_size_img, up_coord, down_coord, color_list)
    coord_list = get_text_coord(ready_img, mosaic_size)
    encode_text = get_encoding_text(color_name)

    ready_img = add_text(ready_img, encode_text, coord_list, mosaic_size)
    # ready_img = add_grid(ready_img, mosaic_size)
    save_img("result", ready_img)
    save_color_table("table", color_name, encode_text)
